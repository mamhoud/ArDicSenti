import nltk
import sys

from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.stem.isri import ISRIStemmer
from nltk.corpus import stopwords

import xlsxwriter
from pandas._testing import isnull
from xlrd import open_workbook
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook



class mechanism (object) :

    types = {'Root': 0, 'Positive': 1, 'Negative': 2, 'Mis': 3}
    data_set = [['Root', 'Positive', 'Negative', 'Mis'], ['حمد', 40, 90, 50], ['وقع', 0, 0, 0], ['لعب', 0, 0, 0],
                ['شرب', 0, 0, 0], ['مشى', 0, 0, 0]]
    def __init__(self):
        self
    def stemer(self,word):
        stem = ISRIStemmer()
        root = stem.stem(word)
        return root


    def add_word(self,iRow, word):
        path = "inter.xlsx"
        wb = load_workbook(path)
        ws = wb.get_sheet_by_name("Sheet1")
        R = ws.cell(row = iRow +1, column = 1)
        P = ws.cell(row=iRow + 1, column = 2)
        N = ws.cell(row=iRow + 1, column = 3)
        M = ws.cell(row=iRow + 1, column = 4)

        R.value = word
        P.value = 0
        N.value = 0
        M.value = 0
        wb.save(path)

    # Create a workbook and add a worksheet.
    def createFile(self):
        path = "inter.xlsx"
        wb = load_workbook(path)
        if not wb :
            Row = 0
            Col = 0
            workbook = xlsxwriter.Workbook('inter.xlsx')
            worksheet = workbook.add_worksheet()

            # Iterate over the data and write it out row by row.
            for row in self.data_set:
                for record in row:
                    worksheet.write(Row, Col, record)
                    Col += 1
                Row += 1
                Col = 0
            workbook.close()
            return  1
        else:
            return 0

    def search (self,word):
        r = 0
        x = 0
        read = open_workbook(r"inter.xlsx")
        for sheet in read.sheets():
            if sheet.name == 'Sheet1':
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        record = sheet.cell(row, col).value
                        if record == word:
                            r = row
                            x = 1
                            break
                if x == 0 :
                    self.add_word(sheet.nrows, word)
                    r = sheet.nrows
                    break

        return r

    def modifyPos(self,iRow, newVal):
        path = "inter.xlsx"
        wb = load_workbook(path)
        ws = wb.get_sheet_by_name("Sheet1")
        c = ws.cell(row = iRow +1, column = self.types['Positive']+ 1)
        c.value = c.value + newVal
        wb.save(path)

    def modifyNeg(self,iRow, newVal):
        path = "inter.xlsx"
        wb = load_workbook(path)
        ws = wb.get_sheet_by_name("Sheet1")
        c = ws.cell(row = iRow +1, column = self.types['Negative']+ 1)
        c.value = c.value+newVal
        wb.save(path)

    def modifyMis(self,iRow, newVal):
        path = "inter.xlsx"
        wb = load_workbook(path)
        ws = wb.get_sheet_by_name("Sheet1")
        c = ws.cell(row = iRow +1, column = self.types['Mis']+ 1)
        c.value = c.value+newVal
        wb.save(path)
    def Summition(self,Positive, Negative , Misleading):
        per_array = []
        Sum = Positive + Negative + Misleading
        if Positive > 0 :
            Per_pos =(Positive/Sum)*100
            per_array.append(Per_pos)
        else :
            per_array.append(0)
        if Negative > 0:
            Per_Neg = (Negative / Sum)*100
            per_array.append(Per_Neg)
        else:
            per_array.append(0)
        if Misleading > 0:
            Per_mis = (Misleading /Sum)*100
            per_array.append(Per_mis)
        else:
            per_array.append(0)
        return  per_array

    def file_generator(self,positive_percentage = 0 , negative_percentage = 0 , misleading_percentage = 0):
        Positive_final = xlsxwriter.Workbook('Positive_final.xlsx')
        worksheet_p = Positive_final.add_worksheet()
        Negative_final = xlsxwriter.Workbook('Negative_final.xlsx')
        worksheet_n = Negative_final.add_worksheet()
        Misleading_final = xlsxwriter.Workbook('Misleading_final.xlsx')
        worksheet_m = Misleading_final.add_worksheet()

        read = open_workbook("inter.xlsx")
        read_pos = open_workbook("Positive_final.xlsx")
        read_neg = open_workbook("Negative_final.xlsx")
        read_mis = open_workbook("Misleading_final.xlsx")

        for sheet in read.sheets():
                if sheet.name == 'Sheet1':
                    for row in range(sheet.nrows):
                        if row > 0:
                            s = sheet.cell(row, 0).value
                            s1 = sheet.cell(row, 1).value
                            s2 = sheet.cell(row, 2).value
                            s3 = sheet.cell(row, 3).value
                            row_array = self.Summition(s1, s2, s3)

                            worksheet_p.write(0, 0, "Root")
                            worksheet_n.write(0, 0, "Root")
                            worksheet_m.write(0, 0, "Root")
                            worksheet_p.write(0,1,"PERCENTAGE")
                            worksheet_n.write(0,1,"PERCENTAGE")
                            worksheet_m.write(0, 1, "PERCENTAGE")

                            if( int(row_array[0]) >= positive_percentage ) and (int(row_array[0])!= 0) and (positive_percentage != 0):

                                worksheet_p.write(row,0,s)
                                worksheet_p.write(row , 1 , s1)

                            else :
                                worksheet_p.write(row,0,"NULL")

                            if (int(row_array[1]) >= negative_percentage ) and (int(row_array[1])!= 0) and (negative_percentage  != 0):

                                worksheet_n.write(row,0,s)
                                worksheet_m.write(row , 1, s2)
                            else:
                                worksheet_n.write(row, 0, "NULL")

                            if (int(row_array[2]) >= misleading_percentage) and ( int(row_array[2])!= 0) and (misleading_percentage != 0):

                                worksheet_m.write(row,0,s)
                                worksheet_m.write(row , 1 , s3)
                            else:
                                worksheet_m.write(row, 0, "NULL")






        Positive_final.close()
        Negative_final.close()
        Misleading_final.close()


    # createFile()

    # s = Summition(45812, 35 ,2)
    # print(s)

    #file_generator(30,0,2)
    # r = stemer("استشرق")
    # s = search(r)
    #
    # modifyNeg(s, 1)
    #
    # print(r)



    # Write a total using a formula.
    # worksheet.write(row, 0, 'Total')
    # worksheet.write(row, 1, '=SUM(B1:B4)')


